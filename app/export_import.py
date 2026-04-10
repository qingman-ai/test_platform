"""
用例导入导出模块
支持 Excel(.xlsx) 和 YAML(.yaml) 格式
"""
import json
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from sqlalchemy.orm import Session
from . import models

# Excel 列定义（顺序固定）
EXCEL_COLUMNS = [
    ("name", "用例名称"),
    ("method", "请求方法"),
    ("url", "请求URL"),
    ("module", "模块"),
    ("priority", "优先级"),
    ("tags", "标签"),
    ("headers", "请求头(JSON)"),
    ("params", "查询参数(JSON)"),
    ("body", "请求体(JSON)"),
    ("expected_status", "预期状态码"),
    ("assert_keyword", "关键字断言"),
    ("assert_json_field", "JSON字段断言(JSON)"),
    ("assert_max_time", "最大响应时间(秒)"),
    ("extract_vars", "变量提取规则(JSON)"),
]


def export_to_excel(db: Session, module: str = None) -> BytesIO:
    """导出用例为 Excel 文件"""
    # 查询用例
    query = db.query(models.TestCase)
    if module:
        query = query.filter(models.TestCase.module == module)
    cases = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, (field, title) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据
    for row_idx, case in enumerate(cases, 2):
        row_data = {
            "name": case.name,
            "method": case.method,
            "url": case.url,
            "module": case.module or "default",
            "priority": case.priority or 1,
            "tags": case.tags or "",
            "headers": json.dumps(case.headers, ensure_ascii=False) if case.headers else "",
            "params": json.dumps(case.params, ensure_ascii=False) if case.params else "",
            "body": json.dumps(case.body, ensure_ascii=False) if case.body else "",
            "expected_status": case.expected_status or 200,
            "assert_keyword": case.assert_keyword or "",
            "assert_json_field": json.dumps(case.assert_json_field, ensure_ascii=False) if case.assert_json_field else "",
            "assert_max_time": case.assert_max_time or "",
            "extract_vars": json.dumps(case.extract_vars, ensure_ascii=False) if case.extract_vars else "",
        }
        for col_idx, (field, _) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[field])
            cell.border = thin_border

    # 自动调整列宽
    for col_idx, (field, title) in enumerate(EXCEL_COLUMNS, 1):
        max_length = len(title)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_yaml(db: Session, module: str = None) -> str:
    """导出用例为 YAML 字符串"""
    query = db.query(models.TestCase)
    if module:
        query = query.filter(models.TestCase.module == module)
    cases = query.all()

    data = []
    for case in cases:
        item = {
            "name": case.name,
            "method": case.method,
            "url": case.url,
            "module": case.module or "default",
            "priority": case.priority or 1,
            "tags": case.tags or "",
            "headers": case.headers if case.headers else {},
            "params": case.params if case.params else {},
            "body": case.body if case.body else {},
            "expected_status": case.expected_status or 200,
            "assert_keyword": case.assert_keyword or "",
            "assert_json_field": case.assert_json_field if case.assert_json_field else {},
            "assert_max_time": case.assert_max_time or "",
            "extract_vars": case.extract_vars if case.extract_vars else {},
        }
        data.append(item)

    return yaml.dump(data, allow_unicode=True, default_flow_style=False, sort_keys=False)


def _safe_json(value):
    """安全解析 JSON 字符串，解析失败返回空"""
    if not value or value == "":
        return None
    if isinstance(value, dict):
        return value
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return None


def import_from_excel(db: Session, file_bytes: bytes, created_by: str = None) -> dict:
    """从 Excel 导入用例"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active

    # 读取表头，建立列名映射
    header_map = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            # 通过中文表头找到对应的字段名
            for field, title in EXCEL_COLUMNS:
                if cell.value.strip() == title:
                    header_map[col_idx] = field
                    break

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        try:
            row_data = {}
            for col_idx, cell in enumerate(row, 1):
                if col_idx in header_map:
                    row_data[header_map[col_idx]] = cell.value

            # 必填字段校验
            if not row_data.get("name") or not row_data.get("method") or not row_data.get("url"):
                skipped += 1
                continue

            case = models.TestCase(
                name=str(row_data["name"]),
                method=str(row_data["method"]).upper(),
                url=str(row_data["url"]),
                module=str(row_data.get("module", "default") or "default"),
                priority=int(row_data.get("priority", 1) or 1),
                tags=str(row_data.get("tags", "") or ""),
                expected_status=int(row_data.get("expected_status", 200) or 200),
                assert_keyword=str(row_data.get("assert_keyword", "") or "") or None,
                assert_max_time=str(row_data.get("assert_max_time", "") or "") or None,
            )
            case.headers = _safe_json(row_data.get("headers"))
            case.params = _safe_json(row_data.get("params"))
            case.body = _safe_json(row_data.get("body"))
            case.assert_json_field = _safe_json(row_data.get("assert_json_field"))
            case.extract_vars = _safe_json(row_data.get("extract_vars"))

            db.add(case)
            imported += 1

        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")
            skipped += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}


def import_from_yaml(db: Session, yaml_content: str, created_by: str = None) -> dict:
    """从 YAML 导入用例"""
    data = yaml.safe_load(yaml_content)
    if not isinstance(data, list):
        return {"imported": 0, "skipped": 0, "errors": ["YAML格式错误：顶层应为列表"]}

    imported = 0
    skipped = 0
    errors = []

    for idx, item in enumerate(data, 1):
        try:
            if not item.get("name") or not item.get("method") or not item.get("url"):
                skipped += 1
                continue

            case = models.TestCase(
                name=str(item["name"]),
                method=str(item["method"]).upper(),
                url=str(item["url"]),
                module=str(item.get("module", "default") or "default"),
                priority=int(item.get("priority", 1) or 1),
                tags=str(item.get("tags", "") or ""),
                expected_status=int(item.get("expected_status", 200) or 200),
                assert_keyword=str(item.get("assert_keyword", "") or "") or None,
                assert_max_time=str(item.get("assert_max_time", "") or "") or None,
            )

            headers = item.get("headers")
            case.headers = headers if isinstance(headers, dict) else _safe_json(headers)
            params = item.get("params")
            case.params = params if isinstance(params, dict) else _safe_json(params)
            body = item.get("body")
            case.body = body if isinstance(body, dict) else _safe_json(body)
            assert_json = item.get("assert_json_field")
            case.assert_json_field = assert_json if isinstance(assert_json, dict) else _safe_json(assert_json)
            extract = item.get("extract_vars")
            case.extract_vars = extract if isinstance(extract, dict) else _safe_json(extract)

            db.add(case)
            imported += 1

        except Exception as e:
            errors.append(f"第{idx}条: {str(e)}")
            skipped += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}